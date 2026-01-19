"""
Trial Balance to Schedule III Mapping Processor
Implements two-stage mapping with keyword priority logic
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import re
from typing import Dict, List, Tuple, Optional, Any

INPUT_FILE = r'C:\ICAI_Audit_Tool\nw\SRM_Pro\Excel for AI.xlsx'
OUTPUT_SHEET_NAME = 'AM_TBforFS'

def load_workbook_data(file_path: str):
    """Load all required sheets from the workbook"""
    print(f"Loading workbook: {file_path}")
    wb = openpyxl.load_workbook(file_path)
    
    # Load sheets
    tb_sheet = wb['TrialBalance']
    sch3_sheet = wb['Sch3']
    map1_sheet = wb['Map1']
    map2_sheet = wb['Map2']
    
    # Convert sheets to list of dicts
    tb_data = sheet_to_dicts(tb_sheet)
    sch3_data = sheet_to_dicts(sch3_sheet)
    map1_data = sheet_to_dicts(map1_sheet)
    map2_data = sheet_to_dicts(map2_sheet)
    
    print(f"Loaded {len(tb_data)} TrialBalance rows")
    print(f"Loaded {len(sch3_data)} Sch3 rows")
    print(f"Loaded {len(map1_data)} Map1 rows")
    print(f"Loaded {len(map2_data)} Map2 rows")
    
    return wb, tb_data, sch3_data, map1_data, map2_data

def sheet_to_dicts(sheet) -> List[Dict[str, Any]]:
    """Convert sheet to list of dictionaries"""
    headers = []
    for cell in sheet[1]:
        headers.append(cell.value if cell.value else '')
    
    data = []
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        row_dict = {'_row_num': row_idx}
        for idx, value in enumerate(row):
            if idx < len(headers):
                row_dict[headers[idx]] = value
        data.append(row_dict)
    
    return data

def extract_hierarchy_info(tb_row: Dict[str, Any]) -> Dict[str, Any]:
    """
    Extract hierarchy information from TrialBalance row
    Returns: PRIMARY_FIELD_VALUE, PRIOR_TO_PRIMARY_VALUE, hierarchy_values, ledger_name
    """
    # Hierarchy columns in order
    hier_cols = ['Parent', 'Group.$Parent', 'Group.$Parent.1', 'Group.$Parent.2', 
                 'Group.$Parent.3', 'Group.$Parent.4', 'Group.$Parent.5', 
                 'Group.$Parent.6', 'Group.$Parent.7', 'Group.$Parent.8']
    
    hierarchy_values = []
    primary_col_idx = -1
    
    # Build hierarchy list and find Primary
    for idx, col in enumerate(hier_cols):
        value = tb_row.get(col, '')
        if value:
            value_str = str(value).strip()
            hierarchy_values.append(value_str)
            
            # Check if this is the Primary column
            if '_x0004_ Primary' in value_str:
                primary_col_idx = idx
        else:
            hierarchy_values.append('')
    
    # Extract key values
    primary_field_value = ''
    prior_to_primary_value = ''
    
    if primary_col_idx > 0:
        # PRIMARY_FIELD_VALUE = immediate before Primary
        primary_field_value = hierarchy_values[primary_col_idx - 1] if primary_col_idx > 0 else ''
        
        # PRIOR_TO_PRIMARY_VALUE = one before PRIMARY_FIELD_VALUE
        prior_to_primary_value = hierarchy_values[primary_col_idx - 2] if primary_col_idx > 1 else ''
    else:
        # Fallback: use last non-empty hierarchy value
        for val in reversed(hierarchy_values):
            if val and '_x0004_ Primary' not in val:
                primary_field_value = val
                break
    
    ledger_name = str(tb_row.get('Name', '')).strip()
    
    return {
        'primary_field_value': primary_field_value,
        'prior_to_primary_value': prior_to_primary_value,
        'hierarchy_values': hierarchy_values,
        'ledger_name': ledger_name,
        'primary_col_idx': primary_col_idx
    }

def map1_lookup(tally_pre_primary: str, map1_data: List[Dict]) -> str:
    """Lookup Face Note from Map1"""
    tally_pre_primary_lower = tally_pre_primary.lower().strip()
    
    for map1_row in map1_data:
        map1_key = str(map1_row.get('Tally Pre-Primary', '')).lower().strip()
        if map1_key == tally_pre_primary_lower:
            result = map1_row.get('Mapping1 Resultant', '')
            return str(result).strip() if result else 'Not Mapped'
    
    return 'Not Mapped'

def map2_keyword_search_name_only(ledger_name: str, map2_data: List[Dict]) -> Dict[str, Any]:
    """
    Fallback keyword search ONLY on ledger name when all other strategies fail
    Returns: map2_resultant, matched_priority, matched_keyword, multiple_matches, matched_keywords_list, all_candidates
    """
    search_field_lower = ledger_name.lower().strip()
    all_candidate_matches = []
    
    for map2_idx, map2_row in enumerate(map2_data):
        keyword_str = str(map2_row.get('Keyword', '')).strip()
        priority = map2_row.get('Priority', 999)
        
        try:
            priority = int(priority) if priority else 999
        except:
            priority = 999
        
        keywords = [k.strip() for k in keyword_str.split('|') if k.strip()]
        
        for keyword in keywords:
            keyword_lower = keyword.lower()
            pattern = r'\b' + re.escape(keyword_lower) + r'\b'
            if re.search(pattern, search_field_lower):
                all_candidate_matches.append({
                    'map2_resultant': str(map2_row.get('Mapping2 Resultant', '')).strip(),
                    'priority': priority,
                    'keyword': keyword,
                    'keyword_length': len(keyword),
                    'row_order': map2_idx,
                    'field_score': 0,  # High priority since it's name-based
                    'field_type': 'name_fallback'
                })
    
    if not all_candidate_matches:
        return {
            'map2_resultant': 'Not Mapped',
            'matched_priority': None,
            'matched_keyword': '',
            'multiple_matches': False,
            'matched_keywords_list': [],
            'all_candidates': []
        }
    
    # Sort by keyword length, priority, row order
    all_candidate_matches.sort(key=lambda x: (-x['keyword_length'], x['priority'], x['row_order']))
    
    winner = all_candidate_matches[0]
    
    return {
        'map2_resultant': winner['map2_resultant'],
        'matched_priority': winner['priority'],
        'matched_keyword': winner['keyword'] + ' (name)',
        'multiple_matches': False,
        'matched_keywords_list': [],
        'all_candidates': all_candidate_matches
    }

def map2_keyword_search(hier_info: Dict[str, Any], map2_data: List[Dict]) -> Dict[str, Any]:
    """
    Perform Map2 keyword search with priority
    Collects ALL matches from ALL search fields, then sorts globally by specificity
    Returns: map2_resultant, matched_priority, matched_keyword, multiple_matches, matched_keywords_list, all_candidates
    """
    # Build search fields in order
    search_fields = []
    
    # 1. PRIMARY_FIELD_VALUE
    if hier_info['primary_field_value']:
        search_fields.append(('primary', hier_info['primary_field_value']))
    
    # 2. PRIOR_TO_PRIMARY_VALUE
    if hier_info['prior_to_primary_value']:
        search_fields.append(('prior', hier_info['prior_to_primary_value']))
    
    # 3. Remaining hierarchy (exclude PRIMARY and PRIOR already added)
    primary_idx = hier_info['primary_col_idx']
    for idx, val in enumerate(hier_info['hierarchy_values']):
        if val and '_x0004_ Primary' not in val:
            # Skip if already added (PRIMARY or PRIOR)
            if primary_idx >= 0:
                if idx == primary_idx - 1 or idx == primary_idx - 2:
                    continue
            search_fields.append(('hierarchy', val))
    
    # 4. Ledger Name
    if hier_info['ledger_name']:
        search_fields.append(('ledger', hier_info['ledger_name']))
    
    # Collect ALL matches from ALL fields (don't stop at first field)
    all_candidate_matches = []
    
    for field_type, search_field in search_fields:
        search_field_lower = search_field.lower().strip()
        
        for map2_idx, map2_row in enumerate(map2_data):
            keyword_str = str(map2_row.get('Keyword', '')).strip()
            priority = map2_row.get('Priority', 999)
            
            # Try to convert priority to int
            try:
                priority = int(priority) if priority else 999
            except:
                priority = 999
            
            # Split by pipe (|) only
            keywords = [k.strip() for k in keyword_str.split('|') if k.strip()]
            
            for keyword in keywords:
                keyword_lower = keyword.lower()
                
                # Whole word match (case-insensitive) using word boundaries
                pattern = r'\b' + re.escape(keyword_lower) + r'\b'
                if re.search(pattern, search_field_lower):
                    # Field precedence score (primary=0 is best, ledger=3 is worst)
                    field_score = {'primary': 0, 'prior': 1, 'hierarchy': 2, 'ledger': 3}.get(field_type, 4)
                    
                    all_candidate_matches.append({
                        'map2_resultant': str(map2_row.get('Mapping2 Resultant', '')).strip(),
                        'priority': priority,
                        'keyword': keyword,
                        'keyword_length': len(keyword),
                        'row_order': map2_idx,
                        'field_score': field_score,
                        'field_type': field_type
                    })
    
    if not all_candidate_matches:
        # No match found
        return {
            'map2_resultant': 'Not Mapped',
            'matched_priority': None,
            'matched_keyword': '',
            'multiple_matches': False,
            'matched_keywords_list': [],
            'all_candidates': []
        }
    
    # Sort globally by: keyword length (longer first), field_score (primary first), priority (1 first), row_order
    # This ensures "Short-Term Loans and Advances" from ledger name beats "Short-term" from primary field
    all_candidate_matches.sort(key=lambda x: (-x['keyword_length'], x['field_score'], x['priority'], x['row_order']))
    
    winner = all_candidate_matches[0]
    
    # Check for multiple matches at same priority and length
    multiple_matches = False
    matched_keywords_list = [winner['keyword']]
    
    for match in all_candidate_matches[1:]:
        if match['priority'] == winner['priority'] and match['keyword_length'] == winner['keyword_length']:
            multiple_matches = True
            matched_keywords_list.append(match['keyword'])
    
    return {
        'map2_resultant': winner['map2_resultant'],
        'matched_priority': winner['priority'],
        'matched_keyword': winner['keyword'],
        'multiple_matches': multiple_matches,
        'matched_keywords_list': matched_keywords_list,
        'all_candidates': all_candidate_matches  # Return all for fallback
    }

def sch3_resolve_with_fallback(face_note: str, map2_result: Dict[str, Any], sch3_data: List[Dict]) -> Dict[str, Any]:
    """
    Resolve complete hierarchy using Sch3 with aggressive fallback to ALL alternative candidates
    Tries ALL candidates systematically, including those with same keyword but different Map2_Resultant
    Returns: face_item, face_note, subnote, subnote1, subnote2, subnote3, ambiguous, error
    """
    map2_resultant = map2_result['map2_resultant']
    all_candidates = map2_result.get('all_candidates', [])
    
    if face_note == 'Not Mapped':
        return {
            'face_item': 'Not Mapped',
            'face_note': face_note,
            'subnote': 'Not Mapped',
            'subnote1': 'Not Mapped',
            'subnote2': 'Not Mapped',
            'subnote3': 'Not Mapped',
            'ambiguous': False,
            'error': 'Face Note not mapped',
            'used_keyword': ''
        }
    
    # Get all valid SubNote options under this Face Note from Sch3
    valid_subnotes = get_valid_subnotes_for_facenote(face_note, sch3_data)
    
    # STRATEGY 1: Try each candidate in order of specificity
    # This includes ALL Map2 rows that matched, even with same keyword but different resultants
    if all_candidates:
        for idx, candidate in enumerate(all_candidates):
            candidate_resultant = candidate['map2_resultant']
            if candidate_resultant and candidate_resultant != 'Not Mapped':
                result = try_sch3_match(face_note, candidate_resultant, sch3_data)
                if not result['error']:
                    # Success with this candidate
                    result['used_keyword'] = candidate['keyword']
                    if idx > 0:
                        result['used_fallback'] = True  # Mark that we used a fallback
                    return result
    
    # STRATEGY 2: If all direct attempts failed, try matching against valid_subnotes
    # This catches cases where Map2_Resultant is phrased differently but matches Sch3
    if all_candidates and valid_subnotes:
        for candidate in all_candidates:
            candidate_resultant = candidate['map2_resultant']
            if candidate_resultant and candidate_resultant != 'Not Mapped':
                candidate_lower = candidate_resultant.lower().strip()
                # Check if this resultant exists in valid subnotes
                if candidate_lower in valid_subnotes:
                    result = try_sch3_match(face_note, candidate_resultant, sch3_data)
                    if not result['error']:
                        result['used_keyword'] = candidate['keyword']
                        result['used_fallback'] = True
                        return result
    
    # STRATEGY 3: Partial match - try to find any valid subnote that contains the Map2_Resultant
    if all_candidates and valid_subnotes:
        for candidate in all_candidates:
            candidate_resultant = candidate['map2_resultant']
            if candidate_resultant and candidate_resultant != 'Not Mapped':
                candidate_lower = candidate_resultant.lower().strip()
                # Check if any valid subnote contains this resultant (or vice versa)
                for valid_sub in valid_subnotes:
                    if candidate_lower in valid_sub or valid_sub in candidate_lower:
                        # Try the valid subnote from Sch3
                        result = try_sch3_match_by_value(face_note, valid_sub, sch3_data)
                        if not result['error']:
                            result['used_keyword'] = candidate['keyword']
                            result['used_fallback'] = True
                            return result
    
    # All attempts failed - return detailed error
    tried_resultants = [c['map2_resultant'] for c in all_candidates[:3]] if all_candidates else [map2_resultant]
    error_msg = f'Tried {len(all_candidates)} candidates. Primary: "{map2_resultant}". None valid under Face Note "{face_note}"'
    
    return {
        'face_item': 'Not Mapped',
        'face_note': face_note,
        'subnote': 'Not Mapped',
        'subnote1': 'Not Mapped',
        'subnote2': 'Not Mapped',
        'subnote3': 'Not Mapped',
        'ambiguous': False,
        'error': error_msg,
        'used_keyword': map2_result['matched_keyword']
    }

def get_valid_subnotes_for_facenote(face_note: str, sch3_data: List[Dict]) -> set:
    """
    Get all valid SubNote/SubNote1/SubNote2/SubNote3 values for a given Face Note
    """
    face_note_lower = face_note.lower().strip()
    valid_subnotes = set()
    
    for row in sch3_data:
        if str(row.get('Face Note', '')).lower().strip() == face_note_lower:
            for col in ['SubNote', 'SubNote1', 'SubNote2', 'SubNote3']:
                val = str(row.get(col, '')).strip()
                if val:
                    valid_subnotes.add(val.lower().strip())
    
    return valid_subnotes

def try_sch3_match_by_value(face_note: str, subnote_value: str, sch3_data: List[Dict]) -> Dict[str, Any]:
    """
    Try to match a subnote value (from valid_subnotes) in Sch3 under given Face Note
    """
    face_note_lower = face_note.lower().strip()
    subnote_lower = subnote_value.lower().strip()
    
    for sch3_row in sch3_data:
        if str(sch3_row.get('Face Note', '')).lower().strip() != face_note_lower:
            continue
        
        # Check all subnote columns
        for col in ['SubNote', 'SubNote1', 'SubNote2', 'SubNote3']:
            val = str(sch3_row.get(col, '')).lower().strip()
            if val == subnote_lower:
                return {
                    'face_item': str(sch3_row.get('Face Item', '')).strip(),
                    'face_note': str(sch3_row.get('Face Note', '')).strip(),
                    'subnote': str(sch3_row.get('SubNote', '')).strip(),
                    'subnote1': str(sch3_row.get('SubNote1', '')).strip(),
                    'subnote2': str(sch3_row.get('SubNote2', '')).strip(),
                    'subnote3': str(sch3_row.get('SubNote3', '')).strip(),
                    'ambiguous': False,
                    'error': ''
                }
    
    return {
        'face_item': 'Not Mapped',
        'face_note': face_note,
        'subnote': 'Not Mapped',
        'subnote1': 'Not Mapped',
        'subnote2': 'Not Mapped',
        'subnote3': 'Not Mapped',
        'ambiguous': False,
        'error': f'SubNote value "{subnote_value}" not found'
    }

def try_sch3_match(face_note: str, map2_resultant: str, sch3_data: List[Dict]) -> Dict[str, Any]:
    """
    Try to match Map2_Resultant in Sch3 under given Face Note
    """
    # Filter Sch3 by Face Note
    face_note_lower = face_note.lower().strip()
    filtered_sch3 = [
        row for row in sch3_data 
        if str(row.get('Face Note', '')).lower().strip() == face_note_lower
    ]
    
    if not filtered_sch3:
        return {
            'face_item': 'Not Mapped',
            'face_note': face_note,
            'subnote': 'Not Mapped',
            'subnote1': 'Not Mapped',
            'subnote2': 'Not Mapped',
            'subnote3': 'Not Mapped',
            'ambiguous': False,
            'error': f'Face Note "{face_note}" not found in Sch3'
        }
    
    # Try matching Map2_Resultant to SubNote, SubNote1, SubNote2, SubNote3
    map2_lower = map2_resultant.lower().strip()
    matched_rows = []
    
    for sch3_row in filtered_sch3:
        # Check SubNote
        if str(sch3_row.get('SubNote', '')).lower().strip() == map2_lower:
            matched_rows.append(sch3_row)
            continue
        # Check SubNote1
        if str(sch3_row.get('SubNote1', '')).lower().strip() == map2_lower:
            matched_rows.append(sch3_row)
            continue
        # Check SubNote2
        if str(sch3_row.get('SubNote2', '')).lower().strip() == map2_lower:
            matched_rows.append(sch3_row)
            continue
        # Check SubNote3
        if str(sch3_row.get('SubNote3', '')).lower().strip() == map2_lower:
            matched_rows.append(sch3_row)
            continue
    
    if not matched_rows:
        return {
            'face_item': 'Not Mapped',
            'face_note': face_note,
            'subnote': 'Not Mapped',
            'subnote1': 'Not Mapped',
            'subnote2': 'Not Mapped',
            'subnote3': 'Not Mapped',
            'ambiguous': False,
            'error': f'Map2_Resultant "{map2_resultant}" not valid under Face Note "{face_note}"'
        }
    
    # Use first match
    winner = matched_rows[0]
    ambiguous = len(matched_rows) > 1
    
    return {
        'face_item': str(winner.get('Face Item', '')).strip(),
        'face_note': str(winner.get('Face Note', '')).strip(),
        'subnote': str(winner.get('SubNote', '')).strip(),
        'subnote1': str(winner.get('SubNote1', '')).strip(),
        'subnote2': str(winner.get('SubNote2', '')).strip(),
        'subnote3': str(winner.get('SubNote3', '')).strip(),
        'ambiguous': ambiguous,
        'error': ''
    }

def process_trial_balance(tb_data: List[Dict], sch3_data: List[Dict], 
                          map1_data: List[Dict], map2_data: List[Dict]) -> List[Dict]:
    """Main processing function"""
    print("\nProcessing Trial Balance rows...")
    results = []
    
    for idx, tb_row in enumerate(tb_data, 1):
        if idx % 100 == 0:
            print(f"  Processed {idx}/{len(tb_data)} rows...")
        
        # Stage 1: Extract hierarchy
        hier_info = extract_hierarchy_info(tb_row)
        
        # Stage 2: Map1 lookup
        tally_pre_primary = hier_info['primary_field_value']
        face_note_from_map1 = map1_lookup(tally_pre_primary, map1_data)
        
        # Stage 3: Map2 keyword search
        map2_result = map2_keyword_search(hier_info, map2_data)
        
        # Stage 4: Sch3 resolution with fallback
        sch3_result = sch3_resolve_with_fallback(face_note_from_map1, map2_result, sch3_data)
        
        # Stage 5: Ultimate fallback - if still error, try Name column exclusively
        if sch3_result['error'] and hier_info['ledger_name']:
            name_only_result = map2_keyword_search_name_only(hier_info['ledger_name'], map2_data)
            if name_only_result['map2_resultant'] != 'Not Mapped':
                name_sch3_result = sch3_resolve_with_fallback(face_note_from_map1, name_only_result, sch3_data)
                if not name_sch3_result['error']:
                    sch3_result = name_sch3_result
        
        # Determine mapping status
        mapping_status = 'OK'
        mapping_error = sch3_result['error']
        
        if face_note_from_map1 == 'Not Mapped':
            mapping_status = 'ERROR'
            if not mapping_error:
                mapping_error = 'Map1 lookup failed'
        elif map2_result['map2_resultant'] == 'Not Mapped':
            mapping_status = 'PARTIAL'
            if not mapping_error:
                mapping_error = 'Map2 keyword search failed'
        elif sch3_result['error']:
            mapping_status = 'ERROR'
        
        # Build result row (preserve original columns + new columns)
        result_row = tb_row.copy()
        result_row.update({
            'Tally_Pre_Primary': tally_pre_primary,
            'Primary_Field_Value': hier_info['primary_field_value'],
            'Prior_To_Primary_Value': hier_info['prior_to_primary_value'],
            'Map1_FaceNote': face_note_from_map1,
            'Map2_Resultant': map2_result['map2_resultant'],
            'MatchedPriority': map2_result['matched_priority'],
            'MatchedKeyword': map2_result['matched_keyword'],
            'UsedKeyword': sch3_result.get('used_keyword', map2_result['matched_keyword']),
            'MultipleMatches': map2_result['multiple_matches'],
            'AmbiguousSch3Match': sch3_result['ambiguous'],
            'Face Item': sch3_result['face_item'],
            'Face Note': sch3_result['face_note'],
            'SubNote': sch3_result['subnote'],
            'SubNote1': sch3_result['subnote1'],
            'SubNote2': sch3_result['subnote2'],
            'SubNote3': sch3_result['subnote3'],
            'MappingStatus': mapping_status,
            'MappingError': mapping_error
        })
        
        results.append(result_row)
    
    print(f"  Completed processing {len(results)} rows")
    return results

def write_output_sheet(wb, results: List[Dict], tb_sheet_name: str = 'TrialBalance'):
    """Write results to new sheet with formatting"""
    print(f"\nWriting output to sheet: {OUTPUT_SHEET_NAME}")
    
    # Remove sheet if exists
    if OUTPUT_SHEET_NAME in wb.sheetnames:
        del wb[OUTPUT_SHEET_NAME]
    
    # Create new sheet
    ws = wb.create_sheet(OUTPUT_SHEET_NAME)
    
    # Get original TB sheet to preserve column order
    tb_sheet = wb[tb_sheet_name]
    tb_headers = [cell.value for cell in tb_sheet[1]]
    
    # New columns to append
    new_columns = [
        'Tally_Pre_Primary', 'Primary_Field_Value', 'Prior_To_Primary_Value',
        'Map1_FaceNote', 'Map2_Resultant', 'MatchedPriority', 'MatchedKeyword', 'UsedKeyword',
        'MultipleMatches', 'AmbiguousSch3Match',
        'Face Item', 'Face Note', 'SubNote', 'SubNote1', 'SubNote2', 'SubNote3',
        'MappingStatus', 'MappingError'
    ]
    
    all_headers = tb_headers + new_columns
    
    # Write headers
    for col_idx, header in enumerate(all_headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Write data
    for row_idx, result in enumerate(results, 2):
        for col_idx, header in enumerate(all_headers, 1):
            value = result.get(header, '')
            # Convert None to empty string
            if value is None:
                value = ''
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Freeze top row
    ws.freeze_panes = 'A2'
    
    # Auto-fit columns
    for col_idx in range(1, len(all_headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 15
    
    print(f"  Written {len(results)} rows to {OUTPUT_SHEET_NAME}")

def main():
    """Main execution"""
    print("="*80)
    print("Trial Balance to Schedule III Mapping Processor")
    print("="*80)
    
    try:
        # Load data
        wb, tb_data, sch3_data, map1_data, map2_data = load_workbook_data(INPUT_FILE)
        
        # Process
        results = process_trial_balance(tb_data, sch3_data, map1_data, map2_data)
        
        # Write output
        write_output_sheet(wb, results)
        
        # Save workbook
        wb.save(INPUT_FILE)
        print(f"\n✓ SUCCESS: Output written to '{OUTPUT_SHEET_NAME}' in {INPUT_FILE}")
        
        # Print summary
        status_counts = {}
        for result in results:
            status = result.get('MappingStatus', 'UNKNOWN')
            status_counts[status] = status_counts.get(status, 0) + 1
        
        print("\nMapping Summary:")
        for status, count in sorted(status_counts.items()):
            print(f"  {status}: {count} rows ({count/len(results)*100:.1f}%)")
        
    except Exception as e:
        print(f"\n✗ ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        return 1
    
    return 0

if __name__ == '__main__':
    exit(main())
