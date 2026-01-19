import openpyxl

wb = openpyxl.load_workbook('SRM_Pro/Disclosures.xlsx')
ws = wb.active

print(f'Sheet Name: {ws.title}')
print(f'Max Row: {ws.max_row}, Max Col: {ws.max_column}')
print('\n=== Content ===\n')

for row_idx in range(1, ws.max_row + 1):
    row_data = []
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row_idx, col_idx)
        if cell.value:
            row_data.append(f'C{col_idx}: {cell.value}')
    if row_data:
        print(f'Row {row_idx}: {" | ".join(row_data)}')
