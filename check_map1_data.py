import openpyxl

wb = openpyxl.load_workbook(r'C:\ICAI_Audit_Tool\nw\SRM_Pro\Excel for AI.xlsx')
ws = wb['Map1']

print(f"Total rows in sheet: {ws.max_row}")

non_empty = 0
sample_rows = []
for i in range(2, min(ws.max_row + 1, 50)):
    tally_val = ws.cell(i, 3).value  # Column C = Tally Pre-Primary
    mapping_val = ws.cell(i, 4).value  # Column D = Mapping1 Resultant
    if tally_val or mapping_val:
        non_empty += 1
        if len(sample_rows) < 20:
            sample_rows.append(f"Row {i}: Tally='{tally_val}', Mapping='{mapping_val}'")

print(f"\nNon-empty rows (first 50): {non_empty}")
print("\nSample rows:")
for row in sample_rows:
    print(f"  {row}")

# Check if there's data beyond row 20
print(f"\nChecking rows 100-105:")
for i in range(100, 106):
    tally_val = ws.cell(i, 3).value
    if tally_val:
        print(f"  Row {i}: {tally_val}")
