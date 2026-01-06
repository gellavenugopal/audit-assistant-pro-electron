import openpyxl

wb = openpyxl.load_workbook('Trial_Balance_Template (5).xlsx')
ws = wb['Instructions']

print('Instructions Sheet:')
for row in range(1, ws.max_row + 1):
    cell = ws.cell(row=row, column=1)
    if cell.value:
        print(f'{row}: {cell.value}')